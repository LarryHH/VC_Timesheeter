import os
import sys
from datetime import datetime
from tkinter import Tk, Canvas, Text, PhotoImage, Button, WORD
from tkinter.tix import Tree
from tkinterdnd2 import *
import tkinter.font
from pyglet import font

from openpyxl import load_workbook
from form import WindowForm
import webbrowser
import requests

# ⚡ IMPORTANT ⚡ version bump to match release number on github
CURRENT_VERSION = "v1.2.1"

if getattr(sys, 'frozen', False):
    application_path = sys._MEIPASS
else:
    application_path = os.path.dirname(os.path.abspath(__file__))

ASSETS_PATH = os.path.join(os.path.abspath(application_path), 'assets\\')
FONTS_PATH = os.path.join(os.path.abspath(application_path), 'fonts\\')
BUTTON_Y_OFFSET = 36

class WindowDnD():

    @staticmethod
    def relative_to_assets(path: str):
        return os.path.join(ASSETS_PATH, path)

    @staticmethod
    def load_font(path: str):
        font.add_file(os.path.join(FONTS_PATH, path))

    def __init__(self) -> None:
        self.window = Tk()
        self.canvas = Canvas(
            self.window,
            bg="#E2D7D4",
            height=700,
            width=1000,
            bd=0,
            highlightthickness=0,
            relief="ridge"
        )
        self.canvas.place(x=0, y=0)
        self.bg_left = PhotoImage(
            file=WindowDnD.relative_to_assets("bg_left.png"))
        self.canvas.create_image(
            200,
            350,
            image=self.bg_left
        )

        self.canvas.create_rectangle(
            400.0,
            0.0,
            1000.0,
            700.0,
            fill="#FBF9F9",
            outline="")

        self.window.title(f'VC Timesheeter {CURRENT_VERSION}')
        self.window.geometry("1000x700")
        self.window.configure(bg="#E2D7D4")

        WindowDnD.load_font('EngraversGothicBT.ttf')

        self.fill_canvas_right()
        self.fill_canvas_left()
        self.window.resizable(False, False)

    def fill_canvas_left(self):
        self.canvas.create_text(
            84.0,
            116.0,
            anchor="nw",
            text="VIM&CO",
            fill="black",
            font=("EngraversGothic BT", 70 * -1)
        )
        self.canvas.create_text(
            84.0,
            186.0,
            anchor="nw",
            text="Timesheeter",
            fill="#464749",
            font=("EngraversGothic BT", 45 * -1)
        )

        self.canvas.create_text(
            34.0,
            304.0,
            anchor="nw",
            text="Produces a full, blank set of timesheets",
            fill="#464749",
            font=("Varela Round", 20 * -1)
        )
        self.canvas.create_text(
            34.0,
            340.0,
            anchor="nw",
            text="derived from an Excel template",
            fill="#464749",
            font=("Varela Round", 20 * -1)
        )
        self.canvas.create_text(
            34.0,
            377.0,
            anchor="nw",
            text="for the year of choice.",
            fill="#464749",
            font=("Varela Round", 20 * -1)
        )
        self.canvas.create_text(
            26.0,
            658.0,
            anchor="nw",
            text=f"Larry Huynh, {datetime.now().year}",
            fill="#E2D7D4",
            font=("Varela Round", 16 * -1)
        )

        self.canvas.create_text(
            330.0,
            658.0,
            anchor="nw",
            text=f"{CURRENT_VERSION}",
            fill="#E2D7D4",
            font=("Varela Round", 16 * -1)
        )

        def cb_github_link():
            webbrowser.open_new(r"https://github.com/LarryHH/VC_Timesheeter")
        self.github_link = Button(
            borderwidth=0,
            highlightthickness=0,
            bg='#795c5f',
            activebackground='#795c5f',
            activeforeground='#795c5f',
            command=cb_github_link
        )
        self.github_icon = PhotoImage(file=WindowDnD.relative_to_assets(
            "github_logo.png"))  # make sure to add "/" not "\"
        self.github_link.config(image=self.github_icon)
        self.github_link.place(
            x=165.0,
            y=648.0
        )

    def fill_canvas_right(self):
        def cb_instruction_link():
            webbrowser.open_new(
                r"https://github.com/LarryHH/VC_Timesheeter/blob/master/docs/instructions.md")
        self.instruction_link = Button(
            borderwidth=0,
            highlightthickness=0,
            bg='#fbf9f9',
            activebackground='#fbf9f9',
            activeforeground='#fbf9f9',
            command=cb_instruction_link
        )
        self.instruction_icon = PhotoImage(file=WindowDnD.relative_to_assets(
            "doubts-button_3.png"))  # make sure to add "/" not "\"
        self.instruction_link.config(image=self.instruction_icon)
        self.instruction_link.place(
            x=950.0,
            y=20.0
        )

        self.dnd_img = PhotoImage(
            file=WindowDnD.relative_to_assets("dnd_area.png"))
        self.dnd_bg = self.canvas.create_image(
            700.5,
            394.5,
            image=self.dnd_img
        )
        self.dnd_icon_img = PhotoImage(
            file=WindowDnD.relative_to_assets("dnd_icon.png"))

        self.dnd_area = Text(
            bd=0,
            bg="#FFFFFF",
            highlightthickness=0,
            wrap=WORD,
            font=("Arial Bold", 16 * -1)
        )
        self.dnd_area.place(
            x=500.0,
            y=220.0,
            width=400.0,
            height=350.0
        )
        self.dnd_area.image_create('end', image=self.dnd_icon_img)

        self.dnd_area.drop_target_register(DND_FILES)
        self.dnd_area.dnd_bind('<<Drop>>', self.drop_inside_textbox)
        self.dnd_area.configure(state='disabled')

        self.canvas.create_text(
            546.0,
            158.0,
            anchor="nw",
            text="Drag and drop your Excel template!",
            fill="#464749",
            font=("VarelaRound Regular", 20 * -1)
        )

        try:
            response = requests.get(
                'https://api.github.com/repos/larryhh/VC_Timesheeter/releases/latest').json()
            is_current_version = response["tag_name"] == CURRENT_VERSION
            download_link = response["assets"][0]["browser_download_url"]
        except requests.ConnectionError as e:
            print(e)
            is_current_version = True
        if not is_current_version:
            button_update_available = Button(
                borderwidth=0,
                highlightthickness=0,
                command=lambda: webbrowser.open(download_link),
                relief="flat",
                text='⚡Update available. Click here to download⚡',
                bg='#efc4b8',
                fg='white',
                activebackground='#FDA48B',
                activeforeground='#FFFFFF'
            )
            button_update_available['font'] = tkinter.font.Font(
                family='VarelaRound Regular', size=14)
            button_update_available.place(
                x=475.0,
                y=604.0,
                width=450.0,
                height=50.0
            )

    def report_dnd_error(self, msg):
        self.dnd_area.insert('end', msg)
        self.dnd_area.configure(state='disabled')

    def drop_inside_textbox(self, event):
        self.dnd_area.configure(state='normal')
        self.dnd_area.delete('1.0', 'end')

        files = event.data.split(':')

        if len(files) > 2:
            msg = f'Please only drag n\' drop 1 Excel file.'
            self.dnd_area.insert('end', msg)
            self.dnd_area.configure(state='disabled')
            return

        try:
            filename = ":".join(files).strip("{").strip("}")
            wb = load_workbook(filename=filename)
        except Exception:
            msg = f'The file: {":".join(files)} was not able to be loaded into the drag n\' drop area.\nAre you sure this is an Excel file?'
            self.dnd_area.insert('end', msg)
            self.dnd_area.configure(state='disabled')
            return

        sheets = [s.lower() for s in wb.sheetnames]
        date_sheet_name = ''
        date_cells = []
        date_cell = ''
        if 'date' in sheets:
            index = sheets.index('date')
            date_sheet_name = wb.sheetnames[index]
            for row in wb[date_sheet_name].iter_rows():
                for cell in row:
                    if cell.is_date:
                        date_cells.append(cell)
            # should only be 1 date in date sheet
            if len(date_cells) > 1:
                self.report_dnd_error(
                    f"The '{date_sheet_name}' sheet has multiple cells with dates in them. Please ensure only one exists in the '{date_sheet_name}' sheet.")
                return
            elif len(date_cells) == 0:
                self.report_dnd_error(
                    f"The '{date_sheet_name}' sheet has 0 cells with dates in them. Please ensure one exists in the '{date_sheet_name}' sheet.")
                return
            else:
                date_cell = f'{date_cells[0].column_letter}{date_cells[0].row}'
                self.dnd_area.insert('end', date_cell)
        else:  # date sheet must exist
            self.report_dnd_error(
                f"The Excel file: {':'.join(files)} has no 'Date' sheet. Please make sure one exists in the template.")
            return

        self.dnd_area.configure(state='disabled')
        WindowForm(self.window, ":".join(files),
                   wb, date_sheet_name, date_cell)


if __name__ == "__main__":
    print("Opening VC_Timesheeter. Please wait...")
    dnd = WindowDnD()
    dnd.window.mainloop()
